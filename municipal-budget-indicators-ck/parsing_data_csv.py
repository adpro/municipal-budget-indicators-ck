import openpyxl as xl

from decimal import Decimal
from dataclasses import dataclass, fields

from exceptions import *
from data_model import Year_Input_Data



def filter_input_xlsx_data(filepath):
    output = []
    dataframe = xl.load_workbook(filepath)
    # Define variable to read sheet
    df1 = dataframe.active
    if df1.max_column < 5:
        raise WrongXlsxBudgetFormatError
    # Iterate the loop to read the cell values
    # for row in range(0, df1.max_row):
    df1.delete_cols(4,1)
    df1.delete_cols(1,1)

    for row in range(5, df1.max_row):
        # print(f'"{df1.cell(row, 1).value}", "{df1.cell(row, 2).value}", "{df1.cell(row, 3).value}", "{df1.cell(row, 4).value}"')
        if (str(df1.cell(row, 1).value).strip() == '6330' and 'Konsolidace' in df1.cell(row, 3).value.strip()) \
            or (str(df1.cell(row, 2).value).strip() == '5xxx' and 'Běžné výdaje' in df1.cell(row, 3).value.strip()) \
            or (str(df1.cell(row, 2).value).strip() == '6xxx' and 'Kapitálové výdaje' in df1.cell(row, 3).value.strip()) \
            or str(df1.cell(row, 2).value).strip() == '1xxx' \
            or str(df1.cell(row, 2).value).strip() == '2xxx' \
            or str(df1.cell(row, 2).value).strip() == '3xxx' \
            or str(df1.cell(row, 2).value).strip() == '4xxx' \
            or str(df1.cell(row, 2).value).strip().startswith('41') \
            or str(df1.cell(row, 2).value).strip().startswith('42') \
            or str(df1.cell(row, 1).value).strip().startswith('81') \
            or str(df1.cell(row, 1).value).strip().startswith('82') \
            :
            if (df1.cell(row, 1).value == 6330 and 'Konsolidace' in df1.cell(row, 3).value.strip()) \
                or str(df1.cell(row, 1).value).strip().startswith('81') \
                or str(df1.cell(row, 1).value).strip().startswith('82') \
            :
                output.append([str(df1.cell(row, 1).value), Decimal(df1.cell(row, 4).value)*1000])
                # print(f'{df1.cell(row, 1).value} {df1.cell(row, 4).value*1000}')
            else:
                output.append([str(df1.cell(row, 2).value), Decimal(df1.cell(row, 4).value)*1000])
                # print(f'{df1.cell(row, 2).value} {df1.cell(row, 4).value*1000}')
    # print(output)
    return output

def get_FIN_item(input, item) -> Decimal:
    items = filter(lambda x: x[0] in item, input)
    values = [Decimal(x[1]) for x in items]
    return Decimal(sum(values))

def fill_xls_data_with_FIN(filepath, data):
    # print(f'XLS file: >>>{filepath}<<<')
    filtered_input = filter_input_xlsx_data(filepath)
    data.fin_4010 = get_FIN_item(filtered_input, '1xxx')
    data.fin_4020 = get_FIN_item(filtered_input, '2xxx')
    data.fin_4030 = get_FIN_item(filtered_input, '3xxx')
    data.fin_4040 = get_FIN_item(filtered_input, '4xxx')
    data.fin_4210 = get_FIN_item(filtered_input, '5xxx')
    data.fin_4220 = get_FIN_item(filtered_input, '6xxx')
    data.fin_4250 = get_FIN_item(filtered_input, '6330')
    data.fin_4200 = data.fin_4010 + data.fin_4020 + data.fin_4030 + data.fin_4040
    data.fin_4430 = data.fin_4210 + data.fin_4220 - data.fin_4250
    # data.fin_5141 = get_FIN_item(filtered_input,'5141') # placene uroky

    # transfer
    data.fin_4111 = get_FIN_item(filtered_input,'4111')
    data.fin_4112 = get_FIN_item(filtered_input,'4112')
    data.fin_4113 = get_FIN_item(filtered_input,'4113')    # added later
    data.fin_4116 = get_FIN_item(filtered_input,'4116')
    data.fin_4119 = get_FIN_item(filtered_input,'4119')
    data.fin_4121 = get_FIN_item(filtered_input,'4121')
    data.fin_4122 = get_FIN_item(filtered_input,'4122')
    data.fin_4123 = get_FIN_item(filtered_input,'4123')
    data.fin_4129 = get_FIN_item(filtered_input,'4129')
    data.fin_4151 = get_FIN_item(filtered_input,'4151')
    data.fin_4152 = get_FIN_item(filtered_input,'4152')
    data.fin_4153 = get_FIN_item(filtered_input,'4153')
    data.fin_4155 = get_FIN_item(filtered_input,'4155')
    data.fin_4156 = get_FIN_item(filtered_input,'4156')
    data.fin_4159 = get_FIN_item(filtered_input,'4159')
    data.fin_4160 = get_FIN_item(filtered_input,'4160')
    data.fin_4211 = get_FIN_item(filtered_input,'4211')
    data.fin_4212 = get_FIN_item(filtered_input,'4212')
    data.fin_4213 = get_FIN_item(filtered_input,'4213')
    data.fin_4214 = get_FIN_item(filtered_input,'4214')
    data.fin_4216 = get_FIN_item(filtered_input,'4216')
    data.fin_4218 = get_FIN_item(filtered_input,'4218')
    data.fin_4219 = get_FIN_item(filtered_input,'4219')
    data.fin_4221 = get_FIN_item(filtered_input,'4221')
    data.fin_4222 = get_FIN_item(filtered_input,'4222')
    data.fin_4229 = get_FIN_item(filtered_input,'4229')
    data.fin_4231 = get_FIN_item(filtered_input,'4231')
    data.fin_4232 = get_FIN_item(filtered_input,'4232')
    data.fin_4233 = get_FIN_item(filtered_input,'4233')
    data.fin_4234 = get_FIN_item(filtered_input,'4234')
    data.fin_4235 = get_FIN_item(filtered_input,'4235')
    # finance
    data.fin_8122 = get_FIN_item(filtered_input,'8122')
    data.fin_8124 = get_FIN_item(filtered_input,'8124')
    data.fin_8xx2 = get_FIN_item(filtered_input,['8112','8122','8212','8222'])
    data.fin_8x14 = get_FIN_item(filtered_input,['8114','8214'])
    data.fin_8x13 = get_FIN_item(filtered_input,['8113','8213'])
    data.fin_8x24 = get_FIN_item(filtered_input,['8124','8224'])
    data.fin_8x23 = get_FIN_item(filtered_input,['8123','8223'])
    data.fin_8xx4 = get_FIN_item(filtered_input,['8111','8124','8214','8224'])
    # leasing
    # no use

    # internal transfer
    # no use








